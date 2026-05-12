#!/usr/bin/env python3
"""Export final_contacts_scored.csv to xlsx with a README sheet.

Two sheets:
  - README: how to read the file, scoring model, bucket meanings, source pipeline
  - Contacts: the 344 scored contacts (same as CSV)
"""

import csv
import sys
import textwrap
from pathlib import Path

sys.path.insert(0, "/tmp/pylibs")


def hardwrap(text: str, width: int = 100) -> str:
    """Insert explicit newlines so Excel reliably shows multi-line content."""
    if not text:
        return text
    out = []
    for paragraph in text.split("\n"):
        if len(paragraph) <= width:
            out.append(paragraph)
        else:
            out.extend(
                textwrap.wrap(paragraph, width=width, break_long_words=False) or [""]
            )
    return "\n".join(out)


from openpyxl import Workbook  # noqa: E402
from openpyxl.styles import Alignment, Font, PatternFill  # noqa: E402
from openpyxl.utils import get_column_letter  # noqa: E402

BASE = Path("/Users/user/vivica-outreach/source-lists/segments")
SRC = BASE / "final_contacts_scored.csv"
DST = BASE / "final_contacts_scored.xlsx"

BUCKET_FILL = {
    "HOT": PatternFill("solid", fgColor="FFD6D6"),
    "WARM": PatternFill("solid", fgColor="FFE9C2"),
    "COOL": PatternFill("solid", fgColor="DDEEFF"),
    "COLD": PatternFill("solid", fgColor="E0E0E0"),
}


def add_readme(ws):
    bold = Font(bold=True, size=12)
    h1 = Font(bold=True, size=16)
    h2 = Font(bold=True, size=13)
    wrap = Alignment(wrap_text=True, vertical="top")
    header_fill = PatternFill("solid", fgColor="F0F0F0")

    rows = [
        ("Vivica Outreach — Scored Contacts", h1),
        ("", None),
        ("What this document is", h2),
        (
            "Final prioritized list of 344 verified contacts at US reference labs. "
            "Each contact is scored on 6 signals and assigned to a HOT / WARM / "
            "COOL / COLD bucket. The 'Contacts' sheet holds the data, sorted by "
            "vivica_score descending.",
            wrap,
        ),
        ("", None),
        ("Pool", bold),
        (
            "Our CMS-enrichment funnel (NOT to be confused with Petr's HOT universe). "
            "Source: CMS CLIA Q1 2026 (1,849 reference labs) -> Apollo people search "
            "-> Apollo Reveal by ID -> FindyMail SMTP double-verify. "
            "344 unique contacts across 270 companies.",
            wrap,
        ),
        ("", None),
        ("How it was assembled — funnel", h2),
        ("1. CMS CLIA Q1 2026 -> 1,849 reference labs", None),
        ("2. Domain enrichment (Clay + Exa) -> 1,019 unique domains", None),
        (
            "3. Apollo people search -> 1,252 contacts across 383 domains (37% domain hit)",
            None,
        ),
        ("4. Apollo Reveal (email by ID) -> 1,033 verified", None),
        ("5. FindyMail SMTP verify -> 878 valid (14% bounce filtered out)", None),
        ("6. Dedup by email -> 344 unique contacts", None),
        (
            "Full funnel with conversions: source-lists/enrichment-runs/2026-05-11_reference-1849_findymail-email/analytics.md",
            None,
        ),
        ("", None),
        ("Scoring model (max ~120 points)", h2),
        (
            "Calibrated to the actual signal distribution in the data. CLIA age "
            "and certificate type were dropped — on the Q1 2026 cohort there is "
            "no variance (all <=6 months old, 99% Compliance).",
            wrap,
        ),
        ("", None),
        ("Signal | Source | Weights", bold),
        (
            "Facility type | cms_facility_type_name | independent +30 / public_health +15 / non-lab -30",
            None,
        ),
        (
            "    What is 'independent lab'? CMS CLIA facility code 15 — a stand-alone commercial laboratory operating as its own business, NOT a hospital lab, physician office lab (POL), or public agency. Examples: LabCorp, Quest, Cascade Pathology, Regeneron Clinical Genomics. Sells testing as a service.",
            wrap,
        ),
        (
            "    Why independent labs are Vivica's primary ICP: (1) they buy LIMS themselves — no hospital IT forcing Epic Beaker, (2) LIMS pain hits their P&L directly (revenue engine, not cost center), (3) short sales cycle — CEO/owner can decide, (4) cloud multi-tenant SaaS fits their operating model.",
            wrap,
        ),
        (
            "    'public_health_lab' (+15) = state/county/CDC labs. ICP-adjacent but slower procurement. 'non-lab' (-30) = ambulance, mobile_lab, ASC, blood_bank, hospice, tissue_bank — wrongly classified as REFERENCE in CMS data, won't buy a LIMS.",
            wrap,
        ),
        (
            "Test volume | test_volume | 1k-50k +25 / 50k-500k +15 / 1-999 +10 / 500k+ 0 / 0 -10",
            None,
        ),
        (
            "    What is 'test volume'? Annual number of lab tests the lab self-reports to CMS during CLIA certification (drives the CLIA fee tier). Not revenue — literal test count per year. Range in our dataset: 0 to ~11M, median ~10k.",
            wrap,
        ),
        (
            "    Why the 1k-50k band scores highest: small enough that legacy LIMS pain is real and budget decisions are fast, large enough to afford Vivica. <1k = micro/just-launched, no budget. 500k+ = enterprise, multi-year sales cycle. 0 = data missing, usually non-lab facility.",
            wrap,
        ),
        ("Site count | site_count | 0-1 +10 / 2-5 0 / 6+ -10", None),
        ("Persona | persona | CEO/Owner +20 / Lab Dir +15 / Med Dir +10", None),
        ("LinkedIn | contact_linkedin not empty | +5", None),
        (
            "Petr's tier | tier (only if present in his universe) | S+ +30 / S +20 / A +10 / B +5 / D/E -50",
            None,
        ),
        ("", None),
        ("Bucket thresholds — why these cutoffs", h2),
        (
            "Thresholds are anchored to the 'reference contact' baseline. A typical "
            "independent lab + single site + LinkedIn already nets 45 points before "
            "persona or volume — so any meaningful prospect lands at 50+. The cutoffs "
            "express WHAT is missing, not just a number.",
            wrap,
        ),
        ("", None),
        (
            "HOT (>=75)  — 188 contacts (54.7%)  — Independent lab + decision-maker + reasonable volume. Fundamentally a good fit. Wave 1 in SmartLead.",
            wrap,
        ),
        (
            "WARM (50-74) —  97 contacts (28.2%) — Independent lab but one signal is off: non-CEO persona, or 0 volume data, or large enterprise. Still worth the touch. Main pool.",
            wrap,
        ),
        (
            "COOL (25-49) —  26 contacts (7.6%)  — Multiple weak signals: small chain (2-5 sites) OR public-health lab OR low-seniority persona. Top-up wave if quota is open.",
            wrap,
        ),
        (
            "COLD (<25)   —  33 contacts (9.6%)  — Facility type is wrong: ambulance / mobile_lab / ASC / blood_bank / hospice / tissue_bank. These are NOT lab ICP for Vivica — they slipped into REFERENCE bucket by CMS misclassification. SKIP.",
            wrap,
        ),
        ("", None),
        (
            "Typical score arithmetic for a HOT contact: independent (+30) + sweet-spot volume 1k-50k (+25) + single site (+10) + CEO/Owner (+20) + LinkedIn (+5) = 90. Add Petr's S/S+ tier and you're at 110-120.",
            wrap,
        ),
        (
            "Typical COLD: non-lab facility (-30) + no volume (-10) + 2 sites (0) + persona (+10..20) + LinkedIn (+5) = -5 to -15.",
            wrap,
        ),
        ("", None),
        ("Columns on the Contacts sheet", h2),
        ("Base (from verify pipeline):", bold),
        ("src_name — lab company name (from CMS CLIA)", None),
        ("src_domain — primary domain", None),
        ("src_clia — CLIA number (join key for CMS data)", None),
        (
            "persona — Apollo classification: ceo_owner / lab_director / medical_director",
            None,
        ),
        (
            "contact_first_name / contact_full_name / contact_email / contact_title / contact_linkedin",
            None,
        ),
        ("source — typically apollo+findymail_verified", None),
        ("", None),
        (
            "Petr's scoring (from his universe — only 11/344 contacts have it):",
            bold,
        ),
        ("tier — S+/S/A/B/C/D/E or — (not in universe)", None),
        ("score / cohort / sources / primary_reason — external scoring fields", None),
        ("", None),
        ("Our Vivica-fit scoring:", bold),
        ("vivica_score — total score (max ~120)", None),
        ("vivica_bucket — HOT / WARM / COOL / COLD", None),
        (
            "s_facility_type / s_test_volume / s_site_count / s_persona / s_has_linkedin / s_petr_tier — per-signal contribution",
            None,
        ),
        (
            "facility_type_raw / test_volume_raw / site_count_raw — raw values (for transparency)",
            None,
        ),
        ("", None),
        ("How to use", h2),
        ("1. Load Wave 1 into SmartLead: filter vivica_bucket = HOT (188)", None),
        ("2. After Wave 1 -> Wave 2: vivica_bucket = WARM (97)", None),
        ("3. COOL (26) — top-up if quota is not filled", None),
        (
            "4. COLD (33) — DO NOT load. These are ambulances / mobile units, not ICP for Vivica LIMS.",
            None,
        ),
        (
            "5. Blocklist (Russian-speaking): subtract source-lists/segments/russian_confirmed.csv (19) separately.",
            None,
        ),
        ("", None),
        ("Key caveats", h2),
        (
            "- Apollo email_status=verified is wrong 14.4% of the time. All 344 contacts here passed FindyMail SMTP — safe to load.",
            None,
        ),
        (
            "- 333/344 contacts are outside Petr's universe. This is not a bug, it's a different pool (CMS-only pipeline).",
            None,
        ),
        (
            "- 76% of companies have only one persona reached (full 3-persona ICP coverage is rare).",
            None,
        ),
        (
            "- Outreach angle: 'migrate from your current system' (established operators), NOT 'buy right the first time' (that copy is for Petr's HOT universe).",
            None,
        ),
        ("", None),
        ("Related files", h2),
        (
            "source-lists/segments/README.md — overview of the whole segments folder",
            None,
        ),
        (
            "source-lists/segments/scoring_summary.md — distribution, top-20, breakdown",
            None,
        ),
        ("source-lists/segments/score_contacts.py — the scoring script", None),
        ("source-lists/segments/tier_summary.md — Petr's tier distribution", None),
        (
            "source-lists/enrichment-runs/2026-05-11_reference-1849_findymail-email/manifest.md — enrichment run manifest",
            None,
        ),
        ("tracking/data-log.md — log of all data operations", None),
        ("", None),
        ("Scoring run-id: segments-internal-scoring (2026-05-12)", None),
    ]

    col_width = 140
    ws.column_dimensions["A"].width = col_width

    for i, (text, style) in enumerate(rows, start=1):
        cell = ws.cell(row=i, column=1, value=text)
        cell.alignment = Alignment(wrap_text=True, vertical="top")
        if style:
            if style is bold:
                cell.font = bold
            elif style is h1:
                cell.font = h1
            elif style is h2:
                cell.font = h2
                cell.fill = header_fill

        # Compute row height from wrapped text length (≈ chars per line × 15pt)
        if text:
            chars_per_line = int(col_width * 1.0)
            lines = max(1, -(-len(text) // chars_per_line))
            ws.row_dimensions[i].height = max(16, lines * 16)


def add_contacts(ws):
    with open(SRC, newline="", encoding="utf-8") as f:
        reader = csv.reader(f)
        headers = next(reader)
        rows = list(reader)

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="333333")
    for j, h in enumerate(headers, start=1):
        c = ws.cell(row=1, column=j, value=h)
        c.font = header_font
        c.fill = header_fill

    bucket_idx = headers.index("vivica_bucket")
    score_idx = headers.index("vivica_score")
    rationale_idx = (
        headers.index("vivica_rationale") if "vivica_rationale" in headers else -1
    )

    for i, row in enumerate(rows, start=2):
        bucket = row[bucket_idx]
        fill = BUCKET_FILL.get(bucket)
        for j, val in enumerate(row, start=1):
            if j - 1 == score_idx:
                try:
                    val = int(val)
                except ValueError:
                    pass
            c = ws.cell(row=i, column=j, value=val)
            c.alignment = Alignment(wrap_text=True, vertical="top")
            if fill and (j - 1) in (bucket_idx, score_idx):
                c.fill = fill
                c.font = Font(bold=True)
        # Row height driven mainly by rationale column (width 90)
        if rationale_idx >= 0:
            rationale_text = row[rationale_idx]
            lines = max(1, -(-len(rationale_text) // 85))
            ws.row_dimensions[i].height = max(16, lines * 15)

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{len(rows) + 1}"

    widths = {
        "src_name": 38,
        "src_domain": 28,
        "src_clia": 14,
        "persona": 17,
        "contact_first_name": 14,
        "contact_full_name": 24,
        "contact_email": 32,
        "contact_title": 36,
        "contact_linkedin": 36,
        "source": 24,
        "tier": 6,
        "score": 8,
        "cohort": 18,
        "sources": 18,
        "primary_reason": 40,
        "vivica_score": 10,
        "vivica_bucket": 11,
        "vivica_rationale": 90,
        "facility_type_raw": 16,
        "test_volume_raw": 12,
        "site_count_raw": 10,
    }
    for j, h in enumerate(headers, start=1):
        w = widths.get(h, 10)
        ws.column_dimensions[get_column_letter(j)].width = w


def add_report(ws):
    """Discussion-ready summary: bucket profiles, examples, edge cases."""
    from collections import Counter

    with open(SRC, newline="", encoding="utf-8") as f:
        rows = list(csv.DictReader(f))

    bold = Font(bold=True, size=11)
    h1 = Font(bold=True, size=16)
    h2 = Font(bold=True, size=13)
    wrap = Alignment(wrap_text=True, vertical="top")
    header_fill = PatternFill("solid", fgColor="F0F0F0")

    line_no = 0
    # Will be set at the end; need them here for height math
    COL_WIDTHS = [14, 38, 18, 36, 32, 80]
    TOTAL_WIDTH = sum(COL_WIDTHS)

    def calc_height(text: str, width: int) -> float:
        if not text:
            return 16
        lines = 0
        for paragraph in text.split("\n"):
            lines += max(1, -(-len(paragraph) // max(1, int(width * 0.95))))
        return max(16, lines * 15)

    def add(text, style=None, fill=None):
        nonlocal line_no
        line_no += 1
        c = ws.cell(row=line_no, column=1, value=text)
        c.alignment = Alignment(wrap_text=True, vertical="top")
        if style:
            c.font = style
        if fill:
            c.fill = fill
        # Merge across all 6 columns so text uses full sheet width
        if text:
            ws.merge_cells(
                start_row=line_no, start_column=1, end_row=line_no, end_column=6
            )
            ws.row_dimensions[line_no].height = calc_height(text, TOTAL_WIDTH)
        return c

    def add_table(headers, data_rows, bucket_for_color=None):
        nonlocal line_no
        line_no += 1
        for j, h in enumerate(headers, start=1):
            c = ws.cell(row=line_no, column=j, value=h)
            c.font = bold
            c.fill = PatternFill("solid", fgColor="DDDDDD")
            c.alignment = Alignment(wrap_text=True, vertical="top")
        for row_data in data_rows:
            line_no += 1
            row_fill = (
                BUCKET_FILL.get(row_data[bucket_for_color])
                if bucket_for_color is not None
                else None
            )
            row_height = 16
            for j, val in enumerate(row_data, start=1):
                if isinstance(val, str) and val.startswith("="):
                    continue
                c = ws.cell(row=line_no, column=j, value=val)
                c.alignment = Alignment(wrap_text=True, vertical="top")
                if row_fill:
                    c.fill = row_fill
                col_width = COL_WIDTHS[j - 1] if j - 1 < len(COL_WIDTHS) else 20
                row_height = max(row_height, calc_height(str(val), col_width))
            ws.row_dimensions[line_no].height = row_height

    add("Vivica Outreach — Lead Selection Report", h1)
    add("")
    add(
        "This report is meant to be read alongside the Contacts sheet during a "
        "review call. It explains who we picked, why, and where we might be wrong "
        "— so we can decide together whether the selection logic needs to change.",
        wrap,
    )
    add("")

    # ── 1. Headline numbers
    add("1. Headline numbers", h2, header_fill)
    add("")
    add(f"Total verified contacts: {len(rows)}")
    add(
        f"Unique companies: {len(set(r['src_clia'] for r in rows))} "
        f"(out of 1,849 CMS reference labs covered = 37.8%)"
    )
    buckets = Counter(r["vivica_bucket"] for r in rows)
    add(
        f"Bucket split: HOT {buckets['HOT']} / WARM {buckets['WARM']} / "
        f"COOL {buckets['COOL']} / COLD {buckets['COLD']}"
    )
    add("")

    # ── 2. Bucket profiles
    add("2. Bucket profiles — what kind of contact is in each", h2, header_fill)
    add("")
    add_table(
        ["Bucket", "Count", "Profile", "Recommended action"],
        [
            (
                "HOT",
                buckets["HOT"],
                "Independent reference lab + decision-maker (usually CEO/Owner) + meaningful test volume (typically 1k-50k tests/year). Single site or small footprint. Vivica's bullseye ICP — established operator running on legacy LIMS, ready for a migration pitch.",
                "Wave 1 in SmartLead. Highest open + reply expectation.",
            ),
            (
                "WARM",
                buckets["WARM"],
                "Mostly independent labs but one signal is suboptimal: non-CEO persona (Lab Director / Medical Director), zero or unknown test volume, or large enterprise (500k+ tests). Still ICP, but worse hook.",
                "Wave 2. Personalize by persona — Lab Director needs operational angle, Medical Director needs compliance angle.",
            ),
            (
                "COOL",
                buckets["COOL"],
                "Mixed signals: small chains (2-5 sites), public-health labs (state/county), or low-seniority contacts at independents. Possible buyers but harder sale cycle.",
                "Top-up wave. Use only if quota is open.",
            ),
            (
                "COLD",
                buckets["COLD"],
                "Facility type is wrong. These are ambulance services, mobile labs, ambulatory surgery centers, blood banks, hospices, or tissue banks — they sit in the REFERENCE bucket because of CMS classification quirks, but they are NOT operating clinical reference labs and won't buy a LIMS.",
                "DO NOT load. Filter out before SmartLead.",
            ),
        ],
        bucket_for_color=0,
    )
    add("")

    # ── 3. Top 10 per bucket
    by_bucket: dict = {"HOT": [], "WARM": [], "COOL": [], "COLD": []}
    for r in rows:
        by_bucket[r["vivica_bucket"]].append(r)

    for b in ["HOT", "WARM", "COOL", "COLD"]:
        add(
            f"3.{['HOT', 'WARM', 'COOL', 'COLD'].index(b) + 1} {b} — examples",
            h2,
            header_fill,
        )
        add("")
        if not by_bucket[b]:
            add("(empty)")
            add("")
            continue
        sample = (
            by_bucket[b][:8]
            if b == "HOT"
            else (
                by_bucket[b][:5] + by_bucket[b][-3:]
                if len(by_bucket[b]) > 8
                else by_bucket[b]
            )
        )
        add_table(
            ["Score", "Lab", "Persona", "Title", "Email", "Why this score"],
            [
                (
                    r["vivica_score"],
                    r["src_name"],
                    r["persona"],
                    r["contact_title"],
                    r["contact_email"],
                    r["vivica_rationale"],
                )
                for r in sample
            ],
        )
        add("")

    # ── 4. Edge cases to manually review
    add("4. Edge cases worth manual review with the client", h2, header_fill)
    add("")
    add(
        "These are contacts where the score is plausible but the underlying data "
        "is unusual. Worth a sanity check before sending.",
        wrap,
    )
    add("")

    # 4a. HOT with non-CEO persona
    add("4a. HOT scores driven by Petr's tier alone (could be over-rated):", bold)
    high_petr = [r for r in by_bucket["HOT"] if int(r["s_petr_tier"]) >= 20]
    add_table(
        ["Lab", "Tier", "Score", "Rationale"],
        [
            (r["src_name"], r["tier"], r["vivica_score"], r["vivica_rationale"])
            for r in high_petr[:10]
        ],
    )
    add("")

    # 4b. HOT contacts where volume is very low
    add("4b. HOT contacts with near-zero test volume (just-launched labs):", bold)
    low_vol_hot = [
        r
        for r in by_bucket["HOT"]
        if r["test_volume_raw"].isdigit() and int(r["test_volume_raw"]) < 500
    ]
    add_table(
        ["Lab", "Volume", "Score", "Rationale"],
        [
            (
                r["src_name"],
                r["test_volume_raw"],
                r["vivica_score"],
                r["vivica_rationale"],
            )
            for r in low_vol_hot[:10]
        ],
    )
    add("")

    # 4c. WARM that's actually 'other' facility type
    add("4c. WARM/COOL contacts at 'other' facility type (manual classify):", bold)
    other_ftype = [
        r
        for r in rows
        if r["facility_type_raw"] == "other" and r["vivica_bucket"] in ("WARM", "COOL")
    ]
    add_table(
        ["Lab", "Volume", "Bucket", "Score", "Rationale"],
        [
            (
                r["src_name"],
                r["test_volume_raw"],
                r["vivica_bucket"],
                r["vivica_score"],
                r["vivica_rationale"],
            )
            for r in other_ftype[:10]
        ],
    )
    add("")

    # ── 5. Cross-check with Petr
    add("5. Sanity check — agreement with Petr's external scoring", h2, header_fill)
    add("")
    add(
        "Only 11 of 344 contacts (3.2%) appear in Petr's HOT universe. Of those, "
        "all 10 with S+/S tier landed in our HOT bucket — meaning our scoring "
        "is consistent with his where they overlap. The remaining 333 are a "
        "different pool (CMS-only) and need to stand on their own merits.",
        wrap,
    )
    add("")

    # Column widths
    ws.column_dimensions["A"].width = 14
    ws.column_dimensions["B"].width = 38
    ws.column_dimensions["C"].width = 18
    ws.column_dimensions["D"].width = 36
    ws.column_dimensions["E"].width = 32
    ws.column_dimensions["F"].width = 80


def main():
    wb = Workbook()

    ws_readme = wb.active
    ws_readme.title = "README"
    add_readme(ws_readme)

    ws_report = wb.create_sheet("Report")
    add_report(ws_report)

    ws_contacts = wb.create_sheet("Contacts")
    add_contacts(ws_contacts)

    wb.save(str(DST))
    print(f"  → {DST.name}")


if __name__ == "__main__":
    main()
